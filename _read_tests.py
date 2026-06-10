"""Read test cases from an Excel workbook for test-runner orchestration.

Usage:
    python _read_tests.py <excel_path>

Prints a single JSON object to stdout with keys:
    excel_path, sheet, headers, skipped, tests, file_checks
"""

from __future__ import annotations

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

TEST_NAME_COLS = ["test name", "test case name", "testcase", "testcase name", "name"]
STEP_COLS = [
    "test step",
    "test steps",
    "step definition",
    "step definitions",
    "steps",
    "test step definition",
]
RESULT_COLS = ["result", "results", "status", "test result"]
REMARK_COLS = ["remark", "remarks", "comments", "failure reason"]


def find_col(headers: dict[str, int], candidates: list[str]) -> int | None:
    for c in candidates:
        if c in headers:
            return headers[c]
    for c in candidates:
        for header, col in headers.items():
            if c in header or header in c:
                return col
    return None


def derive_name(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    return re.sub(r"_+", "_", name)


def read_tests(excel_path: str) -> dict:
    file_checks = {
        "exists": os.path.exists(excel_path),
        "readable": os.access(excel_path, os.R_OK),
        "writable": os.access(excel_path, os.W_OK),
    }
    if not file_checks["exists"]:
        return {"error": f"File not found: {excel_path}", "file_checks": file_checks}

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheet = None
    for name in wb.sheetnames:
        if name.lower() in ("test cases", "tests"):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx

    test_col = find_col(headers, TEST_NAME_COLS)
    step_col = find_col(headers, STEP_COLS)
    result_col = find_col(headers, RESULT_COLS)
    remark_col = find_col(headers, REMARK_COLS)

    if not test_col or not step_col:
        return {
            "error": "missing required columns (test name and/or step definitions)",
            "excel_path": excel_path,
            "sheet": sheet.title,
            "headers": headers,
            "file_checks": file_checks,
        }
    if not result_col or not remark_col:
        return {
            "error": "missing write-back columns (result and/or remark)",
            "excel_path": excel_path,
            "sheet": sheet.title,
            "headers": headers,
            "file_checks": file_checks,
        }

    tests: list[dict] = []
    skipped: list[dict] = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            skipped.append({"row_number": row_idx, "reason": "entire row blank"})
            continue

        test_name = row[test_col - 1] if test_col <= len(row) else None
        steps = row[step_col - 1] if step_col <= len(row) else None

        if not test_name or str(test_name).strip() == "":
            skipped.append({"row_number": row_idx, "reason": "test name empty"})
            continue
        if not steps or str(steps).strip() == "":
            skipped.append({"row_number": row_idx, "reason": "step definition empty"})
            continue

        tests.append(
            {
                "row_number": row_idx,
                "testcase_name": derive_name(test_name),
                "test_name": str(test_name).strip(),
                "step_definitions": str(steps).strip(),
            }
        )

    wb.close()

    return {
        "excel_path": os.path.abspath(excel_path),
        "sheet": sheet.title,
        "headers": headers,
        "columns": {
            "test_name": test_col,
            "steps": step_col,
            "result": result_col,
            "remark": remark_col,
        },
        "file_checks": file_checks,
        "skipped": skipped,
        "tests": tests,
    }


def main() -> int:
    if len(sys.argv) < 2 or not sys.argv[1].strip():
        print("Usage: python _read_tests.py <excel_path>", file=sys.stderr)
        return 2

    result = read_tests(sys.argv[1].strip())
    print(json.dumps(result, ensure_ascii=False, indent=2))

    if "error" in result:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
